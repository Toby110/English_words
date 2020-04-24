#coding=utf-8
import tornado.web
import tornado.ioloop
import asyncio
import openpyxl
import os
from tinydb import TinyDB, Query
import json
settings={
    "template_path":"template",
    "static_path":"static",
    "debug":"true",
}

class IndexHandler(tornado.web.RequestHandler):
    def get(self):
        db=TinyDB("template/words.json")
        # file_path=os.path.join(os.path.dirname(__file__),'template','test.xlsx')
        # wb=openpyxl.load_workbook(file_path,data_only=True)
        # ws=wb.worksheets[0]
        # max_num=ws.max_row
        # for i in range(1,max_num+1):
        #     db.insert(
        #         {'unit':ws.cell(i,1).value, 
        #         'words':ws.cell(i,2).value,
        #         'means':ws.cell(i,3).value,
        #         'music':f'https://fanyi.baidu.com/gettts?lan=en&text={ws.cell(i,2).value}&spd=3&source=web',
        #         })
        
        self.render('index.html',data=db.all())

def make_app():
    return tornado.web.Application([
        (r'/',IndexHandler)
        ],**settings)

if __name__ == '__main__':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    app=make_app()
    app.listen(8000)
    tornado.ioloop.IOLoop.current().start()   